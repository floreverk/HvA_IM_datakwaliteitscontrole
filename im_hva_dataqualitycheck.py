import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
# from datetime import date
from openpyxl.styles import Font
from openpyxl.chart import BarChart3D, Reference, PieChart
from tkinter import filedialog
from tkinter.ttk import *
from tkinter import *

hva_im_app = Tk()
hva_im_app.title('im_hva datakwaliteitscontrole')


def info():
    infowindow = Toplevel()
    infolabel = Label(infowindow, text="De tool neemt een exportlijst uit Adlib en voert een aantal "
                                       "kwaliteitschecks uit. Deze kan je hierna opslaan als Excel op een locatie naar "
                                       "keuze.\n\nDe export uit Adlib moet volgende kolommen bevatten: \n\n"
                                       "instelling.naam\nobjectnummer\nobjectnaam\ntitel\nassociatie.onderwerp\n"
                                       "associatie.periode\nreproductie.referentie\nonderscheidende_kenmerken\n"
                                       "afmeting.eenheid\nafmeting.waarde\n\nAls format wordt gekozen voor een "
                                       "csv met bij instellingen als veldscheiding ';'. ", justify=LEFT)
    infolabel.grid(row=0, column=0)


def openfile():
    file = filedialog.askopenfilename(title="select file")
    return file


def save_file():
    location = filedialog.askdirectory(title="save file")
    blank2 = Label(hva_im_app, text="", bg="#eff0eb")
    blank2.grid(column=1, row=8)
    show_location = Label(hva_im_app, text='de output kan je hier vinden: ' + location, bg="#eff0eb")
    show_location.grid(column=1, row=9)
    return location


def choose():

    choice = [variable.get(), variable2.get()]
    if tuple(choice) == ('hva', ''):
        def start():
            # input csv and directory (to be defined by user)
            a = pd.read_csv(openfile(), delimiter=';')
            # directory = r"R:\Collectie HVA"

            # date format yyyy-mm-dd
            # today = date.today()

            # wijzig kolomnamen
            a = a.rename(columns={'afmeting.waarde': 'afmetingwaarde', 'afmeting.eenheid': 'afmetingeenheid',
                                  'associatie.onderwerp': 'associatieonderwerp',
                                  'reproductie.referentie': 'afbeelding'})

            # drop herbestemde records (HB)
            search = "HB"
            a = a[~a.objectnummer.str.contains('|'.join(search))]

            # data errors

            # 1. missing object_name
            b = pd.isna(a['objectnaam'])
            b = a[b]

            # 2. missing title
            c = pd.isna(a['titel'])
            c = a[c]

            # 3. missing associated_subject
            d = pd.isna(a['associatieonderwerp'])
            d = a[d]

            # 4. missing images
            e = pd.isna(a['afbeelding'])
            e = a[e]

            # all images in directory: existing excelquery
            xls = pd.ExcelFile(r'V:\Depshare AGB Erfgoed\Industriemuseum-Huis van Alijn\5_KENNIS_STRATEGIE\10_PROJECTEN'
                               r'\COGHENT\WP5_Content\Handleidingen\VOLLEDIGE_R_SCHIJF.xlsx')
            df = pd.read_excel(xls, 'COLLECTIE HVA')

            # comparison missing images & images in directory
            f = pd.merge(e, df, left_on='objectnummer', right_on='Objectnummer', how='left')
            f.drop('Bestandsnaam', inplace=True, axis=1)
            f.drop('Extensie', inplace=True, axis=1)
            f.drop('Bestandsgrootte', inplace=True, axis=1)
            f.drop('Recordnummer', inplace=True, axis=1)
            f.drop('Objectnummer', inplace=True, axis=1)

            # 5. missing afmeting
            g = pd.isna(a['afmetingwaarde'])
            g = a[g]

            # 6. afmeting 2D =! mm
            size_1 = "cm"
            h = (a.loc[(a['onderscheidende_kenmerken'] == 'DOCUMENTAIRE COLLECTIE') &
                       (a['onderscheidende_kenmerken'] != 'BEELD')])
            h = h[h.afmetingeenheid.str.contains(size_1).fillna(False)]

            # 7. afmeting 3D =! cm
            size_2 = "mm"
            i = (a.loc[(a['onderscheidende_kenmerken'] == 'OBJECT') & (a['onderscheidende_kenmerken'] != 'TEXTIEL')])
            i = i[i.afmetingeenheid.str.contains(size_2).fillna(False)]

            # afmetingen DB
            object_name_2 = "DB"
            j = (a['objectnummer'].str.startswith(object_name_2, na=True))
            j = a[j]
            k = pd.isna(j['afmetingwaarde'])
            k = j[k]
            m = pd.merge(k, df, left_on='objectnummer', right_on='Objectnummer', how='left')
            m.drop('Recordnummer', inplace=True, axis=1)
            m.drop('Objectnummer', inplace=True, axis=1)

            # 9. wrong institution name
            n = a[a["instelling.naam"] != 'Het Huis van Alijn (Gent)']

            # data stats

            # 1. onderscheidende kenmerken
            textiel = a.loc[a["onderscheidende_kenmerken"] == "TEXTIEL"].count()[0]
            objecten = (a.loc[a["onderscheidende_kenmerken"] == "OBJECT"].count()[0]) + textiel
            documentaire_collectie = a.loc[a["onderscheidende_kenmerken"] == "DOCUMENTAIRE COLLECTIE"].count()[0]
            digitale_collectie = a.loc[a["onderscheidende_kenmerken"] == "DIGITALE COLLECTIE"].count()[0]
            fotocollectie = a.loc[a["onderscheidende_kenmerken"] == "BEELD"].count()[0]
            onderscheidende_kenmerken = [objecten, documentaire_collectie, digitale_collectie, fotocollectie]
            o_k = ['Objecten', 'Digitale Collectie', 'Documentaire Collectie', 'Fotocollectie']

            # 2.
            missing_objectnames = b['objectnummer'].count()
            missing_titles = c['objectnummer'].count()
            missing_assos = d['objectnummer'].count()
            missing_images = e['objectnummer'].count()
            missing_labels = ["objectname", "title", "association", "images"]
            missing = [missing_objectnames, missing_titles, missing_assos, missing_images]

            # 3. datering

            achtiende_eeuw = a[a["associatie.periode"].str.contains("18de eeuw", case=False, na=False)].count()[0]
            negentiende_eeuw = a[a["associatie.periode"].str.contains("19de eeuw", case=False, na=False)].count()[0]
            year1900 = (a['associatie.periode'] == 'jaren 1900').sum()
            year1910 = (a['associatie.periode'] == 'jaren 1910').sum()
            year1920 = (a['associatie.periode'] == 'jaren 1920').sum()
            year1930 = (a['associatie.periode'] == 'jaren 1930').sum()
            year1940 = (a['associatie.periode'] == 'jaren 1940').sum()
            year1950 = (a['associatie.periode'] == 'jaren 1950').sum()
            year1960 = (a['associatie.periode'] == 'jaren 1960').sum()
            year1970 = (a['associatie.periode'] == 'jaren 1970').sum()
            year1980 = (a['associatie.periode'] == 'jaren 1980').sum()
            year1990 = (a['associatie.periode'] == 'jaren 1990').sum()
            year2000 = (a['associatie.periode'] == 'jaren 2000').sum()
            year2010 = (a['associatie.periode'] == 'jaren 2010').sum()
            year2020 = (a['associatie.periode'] == 'jaren 2020').sum()
            labels = ['18de eeuw', '19de eeuw', 'jaren 1900', 'jaren 1910', 'jaren 1920', 'jaren 1930', 'jaren 1940',
                      'jaren 1950', 'jaren 1960', 'jaren 1970', 'jaren 1980', 'jaren 1990', 'jaren 2000',
                      'jaren 2010', 'jaren 2020']
            values = [achtiende_eeuw, negentiende_eeuw, year1900, year1910, year1920, year1930, year1940, year1950,
                      year1960,
                      year1970, year1980, year1990, year2000, year2010, year2020]

            # create output
            wb = Workbook()
            ws = wb.active
            ws.title = 'Info'
            ws['A1'] = "list of sheet tab codes"
            ws.append(['original file', 'original csv provided'])
            ws.append(['#01', 'missing object_name'])
            ws.append(['#02', 'missing title'])
            ws.append(['#03', 'missing associated_subject'])
            ws.append(['#04', 'missing images'])
            ws.append(['#05', 'missing afmetingen'])
            ws.append(['#06', 'afmeting 2D =! mm'])
            ws.append(['#07', 'afmeting 3D =! cm'])
            ws.append(['#08', 'afmeting DB is missing'])
            ws.append(['#09', 'wrong institution name'])
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 4
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['F'].width = 4
            ws['D1'] = 'Datering'
            ws['D18'] = 'Onderscheidende Kenmerken'
            ws['D34'] = 'Ontbrekende Basisregistratie'
            ws['A1'].font = Font(bold=True)
            ws['D1'].font = Font(bold=True)
            ws['D18'].font = Font(bold=True)
            ws['D34'].font = Font(bold=True)

            # add charts
            # datering
            r = 2
            for label in labels:
                ws.cell(row=r, column=4).value = label
                r += 1

            r2 = 2
            for value in values:
                ws.cell(row=r2, column=5).value = value
                r2 += 1

            graph1 = Reference(ws, min_col=5, min_row=2, max_col=5, max_row=16)
            titles1 = Reference(ws, min_col=4, min_row=2, max_col=4, max_row=16)
            chart = BarChart3D()
            ws.add_chart(chart, "G2")
            chart.title = 'Datering'
            chart.y_axis.title = 'aantal'
            chart.x_axis.title = 'Datering'
            chart.add_data(graph1)
            chart.set_categories(titles1)

            # onderscheidende kenmerken
            r3 = 19
            for ok in o_k:
                ws.cell(row=r3, column=4).value = ok
                r3 += 1

            r4 = 19
            for onderken in onderscheidende_kenmerken:
                ws.cell(row=r4, column=5).value = onderken
                r4 += 1

            graph2 = Reference(ws, min_col=5, min_row=19, max_col=5, max_row=22)
            graphlabels = Reference(ws, min_col=4, min_row=19, max_col=4, max_row=22)
            chart2 = PieChart()
            chart2.add_data(graph2)
            chart2.set_categories(graphlabels)
            chart2.title = 'Onderscheidende Kenmerken'
            ws.add_chart(chart2, "G18")

            # missing basisreg
            r5 = 35
            for missingl in missing_labels:
                ws.cell(row=r5, column=4).value = missingl
                r5 += 1

            r6 = 35
            for numbermissing in missing:
                ws.cell(row=r6, column=5).value = numbermissing
                r6 += 1

            graph3 = Reference(ws, min_col=5, min_row=35, max_col=5, max_row=38)
            titles3 = Reference(ws, min_col=4, min_row=35, max_col=4, max_row=38)
            chart3 = BarChart3D()
            chart3.title = 'Ontbrekende Basisregistratie'
            chart3.y_axis.title = 'aantal'
            chart3.x_axis.title = 'veld'
            chart3.add_data(graph3)
            chart3.set_categories(titles3)
            ws.add_chart(chart3, "G34")

            # add sheet / lists
            ws = wb.create_sheet("original file")
            rows = dataframe_to_rows(a, index=False)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            ws = wb.create_sheet("#01")
            rows = dataframe_to_rows(b, index=False)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            ws = wb.create_sheet("#02")
            rows = dataframe_to_rows(c, index=False)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            ws = wb.create_sheet("#03")
            rows = dataframe_to_rows(d, index=False)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            ws = wb.create_sheet("#04")
            rows = dataframe_to_rows(f, index=False)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            ws = wb.create_sheet("#05")
            rows = dataframe_to_rows(g, index=False)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            ws = wb.create_sheet("#06")
            rows = dataframe_to_rows(h, index=False)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            ws = wb.create_sheet("#07")
            rows = dataframe_to_rows(i, index=False)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            ws = wb.create_sheet("#08")
            rows = dataframe_to_rows(m, index=False)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            ws = wb.create_sheet("#09")
            rows = dataframe_to_rows(n, index=False)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            wb.save(save_file() + r"\output.xlsx")

        blank = Label(text="", bg="#eff0eb")
        blank.grid(row=4, column=1)

        _info = Label(text="Klik om een csv op te laden", bg="#eff0eb")
        _info.grid(row=5, column=1)

        blank2 = Label(text="", bg="#eff0eb")
        blank2.grid(row=6, column=1)

        buttonstart = Button(hva_im_app, text="csv opladen", padx=20, pady=10, borderwidth=4, bg="#eacb48",
                             command=start)
        buttonstart.grid(row=7, column=1, columnspan=4)

    else:
        def start():
            # input csv and directory (to be defined by user)
            a = pd.read_csv(openfile(), delimiter=';')

            # date format yyyy-mm-dd
            # today = date.today()

            # wijzig kolomnamen
            a = a.rename(columns={'afmeting.waarde': 'afmetingwaarde', 'afmeting.eenheid': 'afmetingeenheid',
                                  'associatie.onderwerp': 'associatieonderwerp',
                                  'reproductie.referentie': 'afbeelding'})

            # drop herbestemde records (HB)
            # afvoer of herbestemming in afdeling

            # data errors

            # 1. missing object_name
            b = pd.isna(a['objectnaam'])
            b = a[b]

            # 2. missing title
            c = pd.isna(a['titel'])
            c = a[c]

            # 3. missing associated_subject
            d = pd.isna(a['associatieonderwerp'])
            d = a[d]

            # 4. missing images
            e = pd.isna(a['afbeelding'])
            e = a[e]

            # all images in directory: existing excelquery
            xls = pd.ExcelFile(r'V:\Depshare AGB Erfgoed\Industriemuseum-Huis van Alijn\5_KENNIS_STRATEGIE\10_PROJECTEN'
                               r'\COGHENT\WP5_Content\Handleidingen\VOLLEDIGE_R_SCHIJF.xlsx')
            df = pd.read_excel(xls, 'COLLECTIE INDUSTRIEMUSEUM')

            # comparison missing images & images in directory
            f = pd.merge(e, df, left_on='objectnummer', right_on='Objectnummer', how='left')
            f.drop('Bestandsnaam', inplace=True, axis=1)
            f.drop('Extensie', inplace=True, axis=1)
            f.drop('Bestandsgrootte', inplace=True, axis=1)
            f.drop('Recordnummer', inplace=True, axis=1)
            f.drop('Objectnummer', inplace=True, axis=1)

            # 5. missing afmeting
            g = pd.isna(a['afmetingwaarde'])
            g = a[g]

            # 6. wrong institution name
            h = a[a["instelling.naam"] != 'Industriemuseum']

            # data stats

            # 1. objectnummers
            af_im = a[a["objectnummer"].str.startswith('AF', na=True)].count()[0]
            print(af_im)
            dc_im = a[a["objectnummer"].str.startswith('DC', na=True)].count()[0]
            print(dc_im)
            d_im = a[a["objectnummer"].str.startswith('D', na=True)].count()[0] - dc_im
            f_im = a[a["objectnummer"].str.startswith('F', na=True)].count()[0]
            re_im = a[a["objectnummer"].str.startswith('RE', na=True)].count()[0]
            v_im = a[a["objectnummer"].str.startswith('V', na=True)].count()[0]
            print(v_im)
            objectnummer = [af_im, dc_im, d_im, f_im, re_im, v_im]
            o_n = ['AF', 'DC', 'D', 'F', 'RE', 'V']

            # 2.
            missing_objectnames = b['objectnummer'].count()
            missing_titles = c['objectnummer'].count()
            missing_assos = d['objectnummer'].count()
            missing_images = e['objectnummer'].count()
            missing_labels = ["objectname", "title", "association", "images"]
            missing = [missing_objectnames, missing_titles, missing_assos, missing_images]

            # 3. datering

            eerste_achtiende = (a['associatie.periode'] == '1ste helft 18de eeuw').sum()
            tweede_achtiende = (a['associatie.periode'] == '2de helft 18de eeuw').sum()
            eerste_negentiende = (a['associatie.periode'] == '1ste helft 19de eeuw').sum()
            tweede_negentiende = (a['associatie.periode'] == '2de helft 19de eeuw').sum()
            eerste_een = (a['associatie.periode'] == '1ste kwart 20ste eeuw').sum()
            eerste_twee = (a['associatie.periode'] == '2de kwart 20ste eeuw').sum()
            eerste_twintigste = (a['associatie.periode'] == '1ste helft 20ste eeuw').sum() + eerste_twee + eerste_een
            tweede_een = (a['associatie.periode'] == '3de kwart 20ste eeuw').sum()
            tweede_twee = (a['associatie.periode'] == 'eind 20ste eeuw').sum()
            tweede_twintigste = (a['associatie.periode'] == '2de helft 20ste eeuw').sum() + tweede_twee + tweede_een
            eerste_eenentwintigste = (a['associatie.periode'] == '1ste kwart 21ste eeuw').sum()
            eerste_eenentwin = (a['associatie.periode'] == '1ste helft 21ste eeuw').sum() + eerste_eenentwintigste

            labels = ['1ste helft 18de eeuw', '2de helft 18de eeuw', '1ste helft 19de eeuw', '2de helft 19de eeuw',
                      '1ste helft 20ste eeuw', '2de helft 20ste eeuw', '1ste helft 21ste eeuw']
            values = [eerste_achtiende, tweede_achtiende, eerste_negentiende, tweede_negentiende, eerste_twintigste,
                      tweede_twintigste, eerste_eenentwin]

            # create output
            wb = Workbook()
            ws = wb.active
            ws.title = 'Info'
            ws['A1'] = "list of sheet tab codes"
            ws.append(['original file', 'original csv provided'])
            ws.append(['#01', 'missing object_name'])
            ws.append(['#02', 'missing title'])
            ws.append(['#03', 'missing associated_subject'])
            ws.append(['#04', 'missing images'])
            ws.append(['#05', 'missing afmetingen'])
            ws.append(['#06', 'wrong institution name'])
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 4
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['F'].width = 4
            ws['D1'] = 'Datering'
            ws['D18'] = 'Objectnummer'
            ws['D34'] = 'Ontbrekende Basisregistratie'
            ws['A1'].font = Font(bold=True)
            ws['D1'].font = Font(bold=True)
            ws['D18'].font = Font(bold=True)
            ws['D34'].font = Font(bold=True)

            # add charts
            # datering
            r = 2
            for label in labels:
                ws.cell(row=r, column=4).value = label
                r += 1

            r2 = 2
            for value in values:
                ws.cell(row=r2, column=5).value = value
                r2 += 1

            graph1 = Reference(ws, min_col=5, min_row=2, max_col=5, max_row=8)
            titles1 = Reference(ws, min_col=4, min_row=2, max_col=4, max_row=8)
            chart = BarChart3D()
            ws.add_chart(chart, "G2")
            chart.title = 'Datering'
            chart.y_axis.title = 'aantal'
            chart.x_axis.title = 'Datering'
            chart.add_data(graph1)
            chart.set_categories(titles1)

            # objectnummer
            r3 = 19
            for ok in o_n:
                ws.cell(row=r3, column=4).value = ok
                r3 += 1

            r4 = 19
            for objn in objectnummer:
                ws.cell(row=r4, column=5).value = objn
                r4 += 1

            graph2 = Reference(ws, min_col=5, min_row=19, max_col=5, max_row=24)
            graphlabels = Reference(ws, min_col=4, min_row=19, max_col=4, max_row=24)
            chart2 = PieChart()
            chart2.add_data(graph2)
            chart2.set_categories(graphlabels)
            chart2.title = 'Objectnummer'
            ws.add_chart(chart2, "G18")

            # missing basisreg
            r5 = 35
            for missingl in missing_labels:
                ws.cell(row=r5, column=4).value = missingl
                r5 += 1

            r6 = 35
            for numbermissing in missing:
                ws.cell(row=r6, column=5).value = numbermissing
                r6 += 1

            graph3 = Reference(ws, min_col=5, min_row=35, max_col=5, max_row=38)
            titles3 = Reference(ws, min_col=4, min_row=35, max_col=4, max_row=38)
            chart3 = BarChart3D()
            chart3.title = 'Ontbrekende Basisregistratie'
            chart3.y_axis.title = 'aantal'
            chart3.x_axis.title = 'veld'
            chart3.add_data(graph3)
            chart3.set_categories(titles3)
            ws.add_chart(chart3, "G34")

            # add sheet / lists
            ws = wb.create_sheet("original file")
            rows = dataframe_to_rows(a, index=False)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            ws = wb.create_sheet("#01")
            rows = dataframe_to_rows(b, index=False)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            ws = wb.create_sheet("#02")
            rows = dataframe_to_rows(c, index=False)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            ws = wb.create_sheet("#03")
            rows = dataframe_to_rows(d, index=False)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            ws = wb.create_sheet("#04")
            rows = dataframe_to_rows(f, index=False)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            ws = wb.create_sheet("#05")
            rows = dataframe_to_rows(g, index=False)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            ws = wb.create_sheet("#06")
            rows = dataframe_to_rows(h, index=False)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            wb.save(save_file() + r"\output.xlsx")

        blank = Label(text="", bg="#eff0eb")
        blank.grid(row=4, column=1)

        _info = Label(text="Klik om een csv op te laden", bg="#eff0eb")
        _info.grid(row=5, column=1)

        blank2 = Label(text="", bg="#eff0eb")
        blank2.grid(row=6, column=1)

        buttonstart = Button(hva_im_app, text="csv opladen", padx=20, pady=10, borderwidth=4, bg="#eacb48",
                             command=start)
        buttonstart.grid(row=7, column=1, columnspan=4)


hva_im_app.configure(bg="#eff0eb")
hva_im_app.geometry("900x550")

Info = Label(text="Maak je keuze en druk op ok", bg="#eff0eb")
Info.grid(row=0, column=1)

blank6 = Label(text="", bg="#eff0eb")
blank6.grid(row=2, column=1)

variable = StringVar()
variable2 = StringVar()

hva = Checkbutton(hva_im_app, text="Huis van Alijn", variable=variable, onvalue="hva", offvalue="", bg="#eff0eb")
im = Checkbutton(hva_im_app, text="Industriemuseum", variable=variable2, onvalue="im", offvalue="", bg="#eff0eb")

hva.grid(row=1, column=1)
im.grid(row=1, column=2)

startknop = Button(hva_im_app, text="ok", padx=20, pady=10, borderwidth=4, bg="#eacb48", command=choose)
startknop.grid(row=3, column=1)

buttoninfo = Button(hva_im_app, text="Info", borderwidth=4, bg="#c5c1bb", command=info)
buttoninfo.grid(row=1, column=0)

hva_im_app.mainloop()
